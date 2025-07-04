from openpyxl import load_workbook
import os
import glob

grade_map = {
    "A+": 4.5, "A0": 4.0,
    "B+": 3.5, "B0": 3.0,
    "C+": 2.5, "C0": 2.0,
    "D+": 1.5, "D0": 1.0, "D": 1.0, "F": 0.0
    # P, NP는 계산에서 제외
}

def calculate_gpa_from_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    total_points = 0
    total_credits = 0

    row = 5
    while True:
        credit_cell = ws[f"I{row}"]
        grade_cell = ws[f"K{row}"]

        credit = credit_cell.value
        grade = grade_cell.value

        if credit is None or grade is None:
            break

        grade = str(grade).strip().upper()

        # P, NP는 GPA 계산에서 제외
        if grade in ("P", "NP"):
            row += 1
            continue

        grade_point = grade_map.get(grade)
        if grade_point is None:
            print(f"⚠️ 알 수 없는 등급: '{grade}' (row {row}) — 무시됨")
            row += 1
            continue

        total_credits += float(credit)
        total_points += float(credit) * grade_point
        row += 1

    if total_credits == 0:
        return 0.0

    return round(total_points / total_credits, 2)

def generate_gpa_svg(gpa, output_path="gpa_badge.svg", school_name=""):
    percentage = gpa / 4.5
    radius = 50
    circumference = 2 * 3.1416 * radius
    stroke_dashoffset = circumference * (1 - percentage)

    svg = f'''<svg width="200" height="160" viewBox="0 0 200 160" xmlns="http://www.w3.org/2000/svg">
    <defs>
        <linearGradient id="grad" x1="0%" y1="0%" x2="100%" y2="0%">
            <stop offset="0%" stop-color="#4facfe"/>
            <stop offset="100%" stop-color="#00f2fe"/>
        </linearGradient>
    </defs>
    <!-- Background Circle -->
    <circle cx="100" cy="70" r="{radius}" fill="#f0f0f0"/>
    
    <!-- Foreground Gradient Circle -->
    <circle
        cx="100" cy="70" r="{radius}"
        fill="none"
        stroke="url(#grad)"
        stroke-width="10"
        stroke-dasharray="{circumference}"
        stroke-dashoffset="{stroke_dashoffset}"
        transform="rotate(-90 100 70)"
        stroke-linecap="round"
    />
    
    <!-- GPA Label -->
    <text x="100" y="45" text-anchor="middle" font-size="14" fill="#555" font-family="Noto Sans KR, sans-serif">GPA</text>
    
    <!-- GPA Value -->
    <text x="100" y="82" text-anchor="middle" font-size="28" fill="#222" font-family="Noto Sans KR, sans-serif">{gpa:.2f}</text>

    <!-- Max GPA Label -->
    <text x="100" y="100" text-anchor="middle" font-size="13" fill="#888" font-family="Noto Sans KR, sans-serif">/ 4.5</text>
    
    <!-- School Name -->
    <text x="100" y="150" text-anchor="middle" font-size="13" fill="#666" font-family="Noto Sans KR, sans-serif">{school_name}</text>
</svg>'''

    with open(output_path, "w") as f:
        f.write(svg)

def update_readme(gpa):
    with open("README.md", "r") as file:
        lines = file.readlines()
    with open("README.md", "w") as file:
        for line in lines:
            if line.startswith("![GPA]") or line.startswith("<img src="):
                file.write('<img src="gpa_badge.svg" alt="GPA badge">\n')
            else:
                file.write(line)

                

if __name__ == "__main__":
    grade_files = glob.glob("assets/grades*.xlsx")
    if not grade_files:
        print("❌ No grades*.xlsx file found in assets/")
        exit(1)
    filename = grade_files[0]
    school_name = filename.split("/")[-1].removeprefix("grades_").removesuffix(".xlsx").replace("_", " ")
    gpa = calculate_gpa_from_excel(filename)
    generate_gpa_svg(gpa, school_name=school_name)
    update_readme(gpa)